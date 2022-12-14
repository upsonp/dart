from openpyxl import load_workbook
from datetime import datetime

from core import utils
from core import models


def load_salt(mission_id, stream):
    error = []

    file_name = str(stream)

    # remove existing samples and re-load from file.
    models.SaltSample.objects.filter(file=file_name, bottle__event__mission_id=mission_id).delete()

    xls_obj = load_workbook(stream, data_only=True)

    sheet = xls_obj.active

    row_count = 0
    scans = []

    samples = []
    for row in sheet.iter_rows(max_col=13, values_only=True):
        row_count += 1
        try:
            sample_id = row[0]
            bottle_label = row[3]
            date_time = row[4]
            calculated_salinity = row[10]
            comments = row[12]

            if bottle_label and \
                    utils.isint(bottle_label) and calculated_salinity and utils.isfloat(calculated_salinity):
                try:
                    date_time = datetime.strptime((datetime.strftime(date_time, '%Y-%m-%d %H:%M:%S') + " +00:00"),
                                                  '%Y-%m-%d %H:%M:%S %z')
                    bottle = models.Bottle.objects.get(
                        bottle_id=bottle_label,
                        event__sample_id__lte=bottle_label,
                        event__end_sample_id__gte=bottle_label,
                        event__mission_id=mission_id,
                        event__instrument__instrument_type=models.InstrumentType.ctd.value)

                    sample = models.SaltSample(file=file_name, bottle=bottle, sample_id=sample_id,
                                               calculated_salinity=calculated_salinity,
                                               sample_date=date_time, comments=comments)
                    samples.append(sample)
                except models.Bottle.DoesNotExist as e:
                    # if the current_id isn't in the samples array a KeyError is thrown, same thing as if we called
                    # models.Sample.objects.get(sample_id=current_id) and a DoesNotExist error was thrown.
                    err = models.Error(mission_id=mission_id, file_name=file_name,
                                       message=f"Bottle with id {bottle_label} Does not exist",
                                       stack_trace=str(e), line=row_count,
                                       error_code=models.ErrorType.missing_id)
                    err.save()
                    error.append(err)
        except Exception as e:
            err = models.Error(mission_id=mission_id, file_name=file_name,
                               message=f"Unexpected error processing file", stack_trace=str(e), line=row_count)
            err.save()
            error.append(err)

    models.SaltSample.objects.bulk_create(samples)

    return error
