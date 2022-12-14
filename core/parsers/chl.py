from openpyxl import load_workbook

from core import models
from core import utils


def load_chl(mission_id, stream):
    error = []
    samples = []

    file_name = str(stream)

    # remove existing samples and re-load from file.
    models.ChlSample.objects.filter(file=file_name, bottle__event__mission_id=mission_id).delete()

    xls_obj = load_workbook(stream, data_only=True)

    sheet = xls_obj.worksheets[0]

    row_count = 0
    cur_sample = None
    for row in sheet.iter_rows(max_col=14, values_only=True):
        row_count += 1
        try:
            sample = row[0]
            volume = row[1]
            chl = row[8]
            phae = row[9]
            mean_c = row[12]
            mean_p = row[13]

            if (sample or cur_sample) and chl and phae and utils.isfloat(chl) and utils.isfloat(phae):
                try:
                    cur_sample = sample if sample else cur_sample
                    bottle = models.Bottle.objects.get(
                        bottle_id=cur_sample,
                        event__sample_id__lte=cur_sample,
                        event__end_sample_id__gte=cur_sample,
                        event__mission_id=mission_id,
                        event__instrument__instrument_type=models.InstrumentType.ctd.value)

                    sample_count = 1 if sample else 2
                    chl_sample = models.ChlSample(file=file_name, bottle=bottle, sample_order=sample_count,
                                                  chl=chl, phae=phae)
                    samples.append(chl_sample)
                except models.Bottle.DoesNotExist as e:
                    # if the current_id isn't in the samples array a KeyError is thrown, same thing as if we called
                    # models.Sample.objects.get(sample_id=current_id) and a DoesNotExist error was thrown.
                    err = models.Error(mission_id=mission_id, file_name=file_name,
                                       message=f"Bottle with id {cur_sample} Does not exist",
                                       stack_trace=str(e), line=row_count,
                                       error_code=models.ErrorType.missing_id)
                    err.save()
                    error.append(err)

        except Exception as e:
            err = models.Error(mission_id=mission_id, file_name=file_name,
                               message=f"Unexpected error processing file",
                               stack_trace=str(e), line=row_count)
            err.save()
            error.append(err)

    models.ChlSample.objects.bulk_create(samples)

    return error
