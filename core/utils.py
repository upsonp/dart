import os
import re
import threading
import time

import ctd
import io

import pandas
import pytz

from openpyxl import load_workbook

from timeit import default_timer as timer

from datetime import datetime
from os.path import join, isfile

from django.http import JsonResponse

from core import models
from threading import Lock

from . import validations

processing = []


def isfloat(num):
    try:
        float(num)
        return True
    except ValueError:
        return False


def isint(num):
    try:
        int(num)
        return True
    except ValueError:
        return False


def convertDMS_degs(dms_string):
    dms = dms_string.split()
    nsew = dms[2].upper()  # north, south, east, west
    degs = (float(dms[0]) + float(dms[1]) / 60) * (-1 if (nsew == 'S' or nsew == 'W') else 1)

    return degs


def convertDegs_DMS(dd):
    d = int(dd)
    m = float((dd-d)*60.0)

    return [d, m]


def get_files(request):
    type = models.FileType[request.GET['file_type'].lower()] if 'file_type' in request.GET else None
    mission = request.GET['mission_id'] if 'mission_id' in request.GET else None

    files = []
    t_log_dir = models.DataFileDirectory.objects.filter(mission_id=mission, file_types__file_type=type.value)
    if t_log_dir:
        log_dir = t_log_dir[0]
    else:
        return JsonResponse({'files':[]})

    for path in os.listdir(log_dir.directory):
        if os.path.isfile(join(log_dir.directory, path)) and path.lower().endswith(type.label.lower()):
            if len(log_dir.data_files.filter(file=path)) <= 0:
                files.append(models.DataFile(directory=log_dir, file_type=type.value, file=path, processed=False))

    models.DataFile.objects.bulk_create(files)

    resp = JsonResponse({'files': [{'fid': f.pk, 'file_name': f.file.name, 'processed': f.processed,
                                    'errors': len(models.Error.objects.filter(file=f)) > 0} for f in
                                   models.DataFile.objects.filter(directory=log_dir)]})

    return resp

def get_ctd_files(request):
    btl_files = []
    mission_id = request.GET['mission_id']
    if 'fid' in request.GET:

        fid = request.GET["fid"]

        btl_dir = models.DataFileDirectory.objects.get(mission_id=mission_id,
                                                       file_types__file_type=models.FileType.btl.value)
        btl_file = btl_dir.data_files.get(pk=fid)
        btl_file.log_errors.all().delete()

        read_ctd(btl_file)
        btl_files.append(btl_file)

        errors = models.Error.objects.filter(file=btl_file)
        if len(errors) > 0:
            return JsonResponse({'errors': [{"pk": e.pk, "line": e.line, "msg": e.message, "trace": e.stack_trace}
                                            for e in errors]})
        else:
            btl_file.processed = True
            btl_file.save()

    return JsonResponse({'action': 'updated'})


def process_elog(request):
    log_files = []
    mission_id = request.GET['mission_id']
    if 'fid' in request.GET:

        fid = request.GET["fid"]

        log_dir = models.DataFileDirectory.objects.get(mission_id=mission_id,
                                                       file_types__file_type=models.FileType.log.value)
        log_file = log_dir.data_files.get(pk=fid)
        log_file.log_errors.all().delete()

        # events = models.Event.objects.filter(actions__file=log_file).distinct()
        # events.delete()

        read_elog(log_file)

        validations.validate_events(log_file)

        errors = models.Error.objects.filter(file=log_file)
        if len(errors) > 0:
            return JsonResponse({'errors': [{"pk": e.pk, "line": e.line, "msg": e.message, "trace": e.stack_trace}
                                            for e in errors]})
        else:
            log_file.processed = True
            log_file.save()

    return JsonResponse({'action': 'updated'})


def read_elog(log_file):
    print(f"Processing {log_file.file.name}")
    file = join(log_file.directory.directory, log_file.file.name)
    stream = io.StringIO(open(file=file, mode="r").read())

    mission = log_file.directory.mission

    # All mid objects start with $@MID@$ and end with a series of equal signs and a blank line.
    # Using regular expressions we'll split the whole stream in to mid objects, then process each object
    paragraph = re.split('====*\n\n', stream.read().strip())
    t = time.time()
    print(f"start: {(time.time() - t)}")
    mid_map = {'file': log_file, 'buffer':{}}
    for mid in paragraph:
        # Each variable in a mid object starts with the label followed by a colon followed by the value
        tmp = re.findall('(.*?): *(.*?)\n', mid)

        # easily convert the (label, value) tuples into a dictionary
        buffer = dict(tmp)

        # pop off the mid object number used to reference this process if there is an issue processing the mid object
        mid_obj = buffer.pop('$@MID@$')

        # process the buffer creating the event objects
        mid_map['buffer'][mid_obj] = buffer
    print(f"load buffer: {(time.time() - t)}")

    mids = mid_map['buffer'].keys()
    process_stations_instrumnets(log_file, mid_map, mids)
    print(f"stations/instruments: {(time.time() - t)}")

    process_events(log_file, mid_map, mids)
    print(f"events: {(time.time() - t)}")

    process_attachments_actions(log_file, mid_map, mids)
    print(f"attachments/actions: {(time.time() - t)}")

    process_variables(log_file, mid_map, mids)
    print(f"variables: {(time.time() - t)}")
    print(f"end: {(time.time() - t)}")


def process_stations_instrumnets(log_file, mid_map, mids):
    mission = log_file.directory.mission
    buff = mid_map['buffer']

    e_stations = [s.name for s in models.Station.objects.all()]
    e_instruments = [i.name for i in models.Instrument.objects.all()]

    stations = []
    instruments = []
    for m in mids:
        try:
            station = buff[m]['Station']
            instrument = buff[m]['Instrument']

            if station not in e_stations:
                stations.append(models.Station(name=station))

            if instrument not in e_instruments:
                try:
                    inst_type = models.InstrumentType[instrument.lower()].value
                except KeyError:
                    # if an unknown type is recieved consider this an 'other' event
                    inst_type = models.InstrumentType['other'].value

                instruments.append(models.Instrument(name=instrument, instrument_type=inst_type))
        except Exception as e:
            error = models.Error(mission=mission, file=log_file, line=m,
                                 message=f"ERR: {log_file.file.name} processing stations and instruments for $@MID@$: "
                                         f"{m}", stack_trace=str(e))
            error.save()

    models.Station.objects.bulk_create(stations)
    models.Instrument.objects.bulk_create(instruments)


def process_events(log_file, mid_map, mids):
    mission = log_file.directory.mission
    buf = mid_map['buffer']

    e_events = [e.event_id for e in models.Event.objects.filter(mission=mission)]

    p_events = []
    c_events = []
    u_events = {'events': [], 'fields':[]}
    for m in mids:
        try:
            event_id = int(buf[m]["Event"])
            if event_id in p_events:
                # if an event has already been processed, don't process it again
                continue

            p_events.append(event_id)

            # we're done with these objects
            station_name = buf[m].pop('Station')
            instrument_name = buf[m].pop('Instrument')
            sample_id = buf[m].pop('Sample ID')
            sample_id = sample_id if sample_id != "" else None

            end_sample_id = buf[m].pop('End_Sample_ID')
            end_sample_id = end_sample_id if end_sample_id != "" else None

            station = models.get_station(name=station_name)
            instrument = models.get_instrument(instrument_name=instrument_name)

            if event_id not in e_events:
                c_events.append(models.Event(mission=mission, event_id=event_id, station=station,
                                             instrument=instrument, sample_id=sample_id, end_sample_id=end_sample_id))
            else:
                e = models.Event.objects.get(mission=mission, event_id=event_id)
                update = False
                if e.station != station:
                    e.station = station
                    if station not in u_events['fields']:
                        u_events['fields'].append('station')
                    update = True

                if e.instrument != instrument:
                    e.instrument = instrument
                    if instrument not in u_events['fields']:
                        u_events['fields'].append('instrument')
                    update = True

                if e.sample_id != sample_id:
                    e.sample_id = sample_id
                    if sample_id not in u_events['fields']:
                        u_events['fields'].append('sample_id')
                    update = True

                if e.end_sample_id != end_sample_id:
                    e.end_sample_id = end_sample_id
                    if end_sample_id not in u_events['fields']:
                        u_events['fields'].append('end_sample_id')
                    update = True

                if update:
                    u_events['events'].append(e)
        except Exception as e:
            error = models.Error(mission=mission, file=log_file, line=m,
                                 message=f"ERR: {log_file.file.name} processing events for $@MID@$: "
                                         f"{m}", stack_trace=str(e))
            error.save()

    models.Event.objects.bulk_create(c_events)
    if u_events['fields']:
        models.Event.objects.bulk_update(objs=u_events['events'], fields=u_events['fields'])


def process_attachments_actions(log_file, mid_map, mids):
    mission = log_file.directory.mission
    buf = mid_map['buffer']

    instr_sensors = []
    c_actions = []
    u_actions = {'actions':[], 'fields': []}
    cur_event = None
    for m in mids:
        try:
            # This date is the date of the last elog update and isn't accurate for recording purpose
            buf[m].pop('Date')

            event_id = buf[m]['Event']
            event = models.Event.objects.get(mission=mission, event_id=event_id)
            evt_type_t = buf[m]["Action"].lower().replace(' ', '_')

            att_str = buf[m].pop('Attached')
            time_position = buf[m].pop("Time|Position").split(" | ")
            comment = buf[m].pop("Comment")

            if cur_event != event_id:
                atts = att_str.split(" | ")
                for a in atts:
                    if a.strip() != '':
                        if len(models.InstrumentSensor.objects.filter(event=event, name=a)) <= 0:
                            instr_sensors.append(models.InstrumentSensor(event=event, name=a))
        except Exception as e:
            error = models.Error(mission=mission, file=log_file, line=m,
                                 message=f"ERR: {log_file.file.name} processing attachments for $@MID@$: "
                                         f"{m}", stack_trace=str(e))
            error.save()

        try:
            # this is a 'naive' date time with no time zone. But it should always be in UTC
            time = f"{time_position[1][0:2]}:{time_position[1][2:4]}:{time_position[1][4:6]}"
            date_time = datetime.strptime(f"{time_position[0]} {time} +00:00", '%Y-%m-%d %H:%M:%S %z')

            lat = convertDMS_degs(time_position[2])
            lon = convertDMS_degs(time_position[3])

            try:
                evt_type = models.ActionType[evt_type_t].value
            except KeyError:
                # if an unknown type is received consider this an 'other' event
                evt_type = models.ActionType['other'].value

            # if an event already contains this action, we'll update it
            if len(event.actions.filter(action_type=evt_type)) > 0:
                update = False

                if evt_type == models.ActionType.other:
                    action = event.actions.get(action_type=evt_type, action_type_other=evt_type_t)
                else:
                    action = event.actions.get(action_type=evt_type)

                if action.latitude != lat:
                    action.latitude = lat
                    if 'latitude' not in u_actions['fields']:
                        u_actions['fields'].append('latitude')
                    update = True

                if action.longitude != lon:
                    action.longitude = lon
                    if 'longitude' not in u_actions['fields']:
                        u_actions['fields'].append('longitude')
                    update = True

                if action.mid != m:
                    action.mid = m
                    if 'mid' not in u_actions['fields']:
                        u_actions['fields'].append('mid')
                    update = True

                if action.comment != comment:
                    action.comment = comment
                    if 'comment' not in u_actions['fields']:
                        u_actions['fields'].append('comment')
                    update = True

                if update:
                    u_actions['actions'].append(action)

            else:
                action = models.Action(file=log_file, event=event,
                                       date_time=date_time, latitude=lat, longitude=lon,
                                       mid=m, action_type=evt_type, comment=comment)
                if evt_type == models.ActionType.other:
                    action.action_type_other = evt_type_t

                c_actions.append(action)
        except Exception as e:
            error = models.Error(mission=mission, file=log_file, line=m,
                                 message=f"ERR: {log_file.file.name} processing actions for $@MID@$: "
                                         f"{m}", stack_trace=str(e))
            error.save()

        cur_event = event_id

    models.InstrumentSensor.objects.bulk_create(instr_sensors)
    models.Action.objects.bulk_create(c_actions)
    if u_actions['fields']:
        models.Action.objects.bulk_update(objs=u_actions['actions'], fields=u_actions['fields'])


def process_variables(log_file, mid_map, mids):
    mission = log_file.directory.mission
    buf = mid_map['buffer']

    fields = []
    for m in mids:
        try:
            event_id = buf[m].pop('Event')
            event = models.Event.objects.get(mission=mission, event_id=event_id)
            evt_type_t = buf[m].pop("Action").lower().replace(' ', '_')

            try:
                evt_type = models.ActionType[evt_type_t].value
            except KeyError:
                # if an unknown type is received consider this an 'other' event
                evt_type = models.ActionType['other'].value

            if evt_type == models.ActionType.other:
                action = event.actions.get(action_type=evt_type, action_type_other=evt_type_t)
            else:
                action = event.actions.get(action_type=evt_type)

            # models.get_variable_name(name=k) is going to be a bottle neck if a variable doesn't already exist
            for k, v in buf[m].items():
                fields.append(models.VariableField(action=action, name=models.get_variable_name(name=k), value=v))
        except Exception as e:
            error = models.Error(mission=mission, file=log_file, line=m,
                                 message=f"ERR: {log_file.file.name} processing variables for $@MID@$: "
                                         f"{m}", stack_trace=str(e))
            error.save()

    models.VariableField.objects.bulk_create(fields)


def process_ctd(request, mission_id):
    ctd_files = []
    updated = []
    errors = []
    if 'fid' in request.GET:
        fid = request.GET["fid"]

        processing.append(fid)
        ctd_file = models.DataFile.objects.get(pk=fid, file_type=models.FileType.btl.value)

        try:
            read_ctd(ctd_file)
            ctd_files.append(ctd_file)

            post_ctd_validation(ctd_files)
            ctd_file.processed = True
            ctd_file.save()

            updated.append(fid)
        except models.Event.DoesNotExist as e:
            err = models.Error(mission_id=mission_id, file=ctd_file, line=-1, message=f"Error Processing file",
                               stack_trace=str(e))
            err.save()

    elif 'did' in request.GET:
        did = request.GET["did"]

        post_ctd_validation(ctd_files)

    errors = models.Error.objects.filter(file=ctd_file)
    return JsonResponse({'updated': updated, "errors": [{"pk": e.pk, "line": e.line, "msg": e.message, "trace": e.stack_trace}
                                            for e in errors]})


def read_ctd(ctd_file):
    if models.FileType(ctd_file.file_type) == models.FileType.btl:
        read_btl(ctd_file)


def read_btl(btl_file):
    models.Error.objects.filter(file=btl_file).delete()

    data_frame = ctd.from_btl(btl_file.file_path)
    metadata = getattr(data_frame, "_metadata")
    header = metadata['header'].split('\n')
    # lat_str = None
    # lon_str = None
    event_number = None

    for h in header:
        if 'event_number:' in h.lower():
            h_str = h.split(":")
            event_number = h_str[1].strip()

    event = models.Event.objects.get(mission=btl_file.directory.mission,
                                     instrument__instrument_type=models.InstrumentType.ctd.value,
                                     event_id=int(event_number))

    # we only want to use rows in the BTL file marked as 'avg'
    data_frame = data_frame[data_frame['Statistic'] == 'avg']

    col_headers = []
    for c in data_frame.columns:
        col_headers.append(c)

    pop = ['Bottle', 'Bottle_', 'Date', 'Statistic', 'Latitude', 'Longitude']
    for h in pop:
        try:
            b_idx = col_headers.index(h)
            col_headers.pop(b_idx)
        except ValueError:
            # if the label doesn'e exists, which might happen in the case of 'Bottle_' a value error is raised
            pass

    sensors = []
    for h in col_headers:
        # I've found that sensor columns usually have a naming convention where its xxx#yyy where the number denotes
        # the primary (0) sensor and the secondary (1) sensor. However, what follows the number is also relevant
        # to the sensor. Usually what follows the sensor priority denotes the unit. (i.e Sbeox0ML/L vs. Sbeox0V)
        c_name = re.split("(\d)", h, 1)

        priority = int(c_name[1]) + 1 if len(c_name) > 1 else None
        units = c_name[2].lower() if len(c_name) > 2 else None

        sensor = models.Sensor.objects.filter(name=c_name[0])
        if priority:
            sensor = sensor.filter(priority=priority)

        if units:
            sensor = sensor.filter(units=units.lower())

        if len(sensor) <= 0:

            sen = models.Sensor(name=c_name[0])
            sen.sensor_type = models.get_sensor_type(c_name[0])

            if priority:
                sen.priority = priority  # priorty in a sensor name starts counting at zero
            if units:
                sen.units = units

            sensors.append(sen)

    models.Sensor.objects.bulk_create(sensors)

    b_create = []
    b_update = []
    bottle_date = data_frame[["Bottle", "Date"]]
    errors = []
    for row in bottle_date.iterrows():
        bottle_id = row[1]["Bottle"]
        date = row[1]["Date"]

        # assume UTC time if a timezone isn't set
        if not hasattr(date, 'timezone'):
            date = pytz.timezone("UTC").localize(date)

        if len(models.Bottle.objects.filter(event=event, bottle_number=bottle_id)) <= 0:
            models.Bottle.objects.filter(event=event, bottle_number=bottle_id).delete()

        bottle_label = event.sample_id + (bottle_id - 1)

        if event.instrument.instrument_type == models.InstrumentType.ctd.value and bottle_label > event.end_sample_id:
            err = models.Error(mission=btl_file.directory.mission, file=btl_file, line=(metadata["skiprows"]+row[0]),
                               message=f"Warning: Bottle ID ({bottle_label}) for event {event.event_id} is outside the "
                                       f"ID range {event.sample_id} - {event.end_sample_id}",
                               stack_trace="",
                               error_code=models.ErrorType.bad_id)
            errors.append(err)

        if len(models.Bottle.objects.filter(event=event, bottle_number=bottle_id)) <= 0:
            b_create.append(models.Bottle(event=event, bottle_id=bottle_label,
                                       bottle_number=bottle_id, date_time=date))
        else:
            update = False
            b = models.Bottle.objects.get(event=event, bottle_number=bottle_id)
            if b.bottle_id != bottle_label:
                b.bottle_id = bottle_label
                update = True

            if b.bottle_number != bottle_id:
                b.bottle_number = bottle_id
                update = True

            if b.date_time != date:
                b.date_time = date
                update = True

            if update:
                b_update.append(b)

    models.Error.objects.bulk_create(errors)
    models.Bottle.objects.bulk_create(b_create)
    models.Bottle.objects.bulk_update(b_update, fields=['bottle_id', 'bottle_number', 'date_time'])

    data_column_create = []
    data_column_update = []
    for c in col_headers:
        sensor = models.get_sensor_name(c)

        df = data_frame[["Bottle", c]]
        for data in df.iterrows():
            bottle = models.Bottle.objects.get(event=event, bottle_number=data[1]["Bottle"])
            b_data = bottle.bottle_data.all()
            if b_data.filter(sensor=sensor):
                ctd_update = b_data.get(sensor=sensor)
                if ctd_update.value != data[1][c]:
                    ctd_update.value = data[1][c]
                    data_column_update.append(ctd_update)
            else:
                data_column_create.append(models.CTDData(bottle=bottle, sensor=sensor, value=data[1][c]))

    models.CTDData.objects.bulk_create(data_column_create)
    models.CTDData.objects.bulk_update(data_column_update, fields=["value"])


def post_ctd_validation(ctd_files):
    pass


def load_samples(request, pk):
    errors = []
    type = request.GET['type']
    files = request.FILES

    file_names = files.keys()
    for fname in file_names:
        models.Error.objects.filter(file_name=fname).delete()

        print(f"Load Start: {fname}")
        if type == 'salt':
            errors += load_salt(pk, files[fname])
        elif type == 'oxy':
            errors += load_oxy(pk, files[fname])
        elif type == 'chl':
            errors += load_chl(pk, files[fname])
        print(f"Load Finished")

    return JsonResponse({"errors": [{"pk": e.pk, "msg": e.message, "trace": e.stack_trace} for e in errors]})


def load_oxy(mission_id, stream):
    error = []
    file_name = str(stream)
    file = None
    if file_name.lower().endswith(".csv"):
        file = pandas.read_csv(stream, skiprows=9).values
    elif file_name.lower().endswith(".dat"):
        file = pandas.read_csv(stream, skiprows=8).values
    else:
        file = pandas.read_excel(stream, skiprows=8).values

    # remove existing oxygen samples and re-load from file.
    models.OxygenSample.objects.filter(file=file_name, bottle__event__mission_id=mission_id).delete()

    row_count = 0
    oxy_bottles = []
    for row in file:
        row_count += 1
        try:
            sample = row[0]
            bottle = row[1]
            o2 = row[2]
            volume = row[10]
            notes = row[13]

            if sample is None and bottle is None and o2 is None and volume is None:
                break  # no more data, break out of the loop rather than read a few hundred blank lines.

            if sample is not None:
                if bottle is None and o2 is None and volume is None:
                    continue
                elif not isfloat(o2):
                    continue

                sample_id = sample.split("_")

                try:
                    bottle = models.Bottle.objects.get(bottle_id=sample_id[0],
                                                       event__sample_id__lte=sample_id[0],
                                                       event__end_sample_id__gte=sample_id[0],
                                                       event__mission_id=mission_id,
                                                       event__instrument__instrument_type=models.InstrumentType.ctd.value)

                    if sample_id[1] == '1':
                        oxy = models.OxygenSample(file=file_name, bottle=bottle, winkler_1=o2)
                        oxy_bottles.append(oxy)
                    else:
                        oxy_bottles[-1].winkler_2 = o2

                except models.Bottle.DoesNotExist as e:
                    # if the current_id isn't in the samples array a KeyError is thrown, same thing as if we called
                    # models.Sample.objects.get(sample_id=current_id) and a DoesNotExist error was thrown.
                    err = models.Error(mission_id=mission_id, file_name=str(stream),
                                       message=f"Bottle with id {sample_id[0]} Does not exist",
                                       stack_trace=str(e), line=row_count,
                                       error_code=models.ErrorType.missing_id)
                    err.save()
                    error.append(err)

        except Exception as e:
            err = models.Error(mission_id=mission_id, file_name=str(stream),
                               message=f"Unexpected error processing file", stack_trace=str(e),
                               line=row_count)
            err.save()
            error.append(err)

    models.OxygenSample.objects.bulk_create(oxy_bottles)
    return error


def load_salt(mission_id, stream):
    error = []

    file_name = str(stream)

    # remove existing samples and re-load from file.
    models.SaltSample.objects.filter(file=file_name, bottle__event__mission_id=mission_id).delete()

    xls_obj = load_workbook(stream, data_only=True)

    sheet = xls_obj.active

    row_count = 0
    scans = []

    samples = []
    for row in sheet.iter_rows(max_col=13, values_only=True):
        row_count += 1
        try:
            sample_id = row[0]
            bottle_label = row[3]
            date_time = row[4]
            calculated_salinity = row[10]
            comments = row[12]

            if bottle_label and isint(bottle_label) and calculated_salinity and isfloat(calculated_salinity):
                try:
                    date_time = datetime.strptime((datetime.strftime(date_time, '%Y-%m-%d %H:%M:%S') + " +00:00"),
                                      '%Y-%m-%d %H:%M:%S %z')
                    bottle = models.Bottle.objects.get(bottle_id=bottle_label,
                                                       event__sample_id__lte=bottle_label,
                                                       event__end_sample_id__gte=bottle_label,
                                                       event__mission_id=mission_id,
                                                       event__instrument__instrument_type=models.InstrumentType.ctd.value)

                    sample = models.SaltSample(file=file_name, bottle=bottle, sample_id=sample_id,
                                               calculated_salinity=calculated_salinity,
                                               sample_date=date_time, comments=comments)
                    samples.append(sample)
                except models.Bottle.DoesNotExist as e:
                    # if the current_id isn't in the samples array a KeyError is thrown, same thing as if we called
                    # models.Sample.objects.get(sample_id=current_id) and a DoesNotExist error was thrown.
                    err = models.Error(mission_id=mission_id, file_name=file_name,
                                       message=f"Bottle with id {bottle_label} Does not exist",
                                       stack_trace=str(e), line=row_count,
                                       error_code=models.ErrorType.missing_id)
                    err.save()
                    error.append(err)
        except Exception as e:
            err = models.Error(mission_id=mission_id, file_name=file_name,
                               message=f"Unexpected error processing file", stack_trace=str(e), line=row_count)
            err.save()
            error.append(err)

    models.SaltSample.objects.bulk_create(samples)

    return error


def load_chl(mission_id, stream):
    error = []
    samples = []

    file_name = str(stream)

    # remove existing samples and re-load from file.
    models.ChlSample.objects.filter(file=file_name, bottle__event__mission_id=mission_id).delete()

    xls_obj = load_workbook(stream, data_only=True)

    sheet = xls_obj.worksheets[0]

    row_count = 0
    cur_sample = None
    for row in sheet.iter_rows(max_col=14, values_only=True):
        row_count += 1
        try:
            sample = row[0]
            volume = row[1]
            chl = row[8]
            phae = row[9]
            mean_c = row[12]
            mean_p = row[13]

            if (sample or cur_sample) and chl and phae and isfloat(chl) and isfloat(phae):
                try:
                    cur_sample = sample if sample else cur_sample
                    bottle = models.Bottle.objects.get(bottle_id=cur_sample,
                                                       event__sample_id__lte=cur_sample,
                                                       event__end_sample_id__gte=cur_sample,
                                                       event__mission_id=mission_id,
                                                       event__instrument__instrument_type=models.InstrumentType.ctd.value)

                    sample_count = 1 if sample else 2
                    chl_sample = models.ChlSample(file=file_name, bottle=bottle, sample_order=sample_count,
                                                  chl=chl, phae=phae)
                    samples.append(chl_sample)
                except models.Bottle.DoesNotExist as e:
                    # if the current_id isn't in the samples array a KeyError is thrown, same thing as if we called
                    # models.Sample.objects.get(sample_id=current_id) and a DoesNotExist error was thrown.
                    err = models.Error(mission_id=mission_id, file_name=file_name,
                                       message=f"Bottle with id {cur_sample} Does not exist",
                                       stack_trace=str(e), line=row_count,
                                       error_code=models.ErrorType.missing_id)
                    err.save()
                    error.append(err)

        except Exception as e:
            err = models.Error(mission_id=mission_id, file_name=file_name,
                               message=f"Unexpected error processing file",
                               stack_trace=str(e), line=row_count)
            err.save()
            error.append(err)

    models.ChlSample.objects.bulk_create(samples)

    return error


def set_directory(request, mission):
    dir = request.GET['dir']
    f_type = request.GET['type']
    mission = models.Mission.objects.get(pk=mission)

    dfd = models.DataFileDirectory(mission=mission, directory=dir)
    dfd.save()

    dfd_type = models.DataFileDirectoryType(directory=dfd, file_type=models.FileType[f_type].value)
    dfd_type.save()

    return JsonResponse({})
